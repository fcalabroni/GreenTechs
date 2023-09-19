#%%
import pandas as pd

# Create a dictionary with the name of tech and level file with key equal to something.xlsx, column_pos equale to 0 and PV, onshore_wind and offshore_wind as keys

tech = 'Offshore wind'

tech_dict = {   'PV': {'column_pos': 4, 'file_name': 'Photovoltaic.xlsx'},
                'Onshore wind': {'column_pos': 5, 'file_name': 'Onshore wind.xlsx'},
                'Offshore wind': {'column_pos': 6, 'file_name': 'Offshore wind.xlsx'}}

# read the file excel LCI.xlsx as a dictionary of dataframes, one for each sheet
df = pd.read_excel(tech_dict[tech]['file_name'], sheet_name=None)

# %% assign the keys of df to a list but drop all the keys that starts with "_"
keys = list(df.keys())
keys = [x for x in keys if not x.startswith('_')]
# %%
# I want to query df['Subcomponent'] for all the rows where Subcomponent == 'Tower' and get the value of column "Weight"
# df['Subcomponent'].query('Subcomponent == "Tower"')['Weight']
 

levels = ['Factor of production','Commodity']
highest = 'Technology'
input = {}
res = {key: {} for key in tech_dict.keys()}
#%%
for t in tech_dict.keys():

    df = pd.read_excel(tech_dict[t]['file_name'], sheet_name=None)

    for level in levels:
        lowest = level

        add_col = {1: {'below': lowest,    'above': 'Subcomponent',  'col_name':'wCOMMinSUBC'}, 
                2: {'below': 'Subcomponent', 'above': 'Component',     'col_name':'wSUBCinCOMP'},
                3: {'below': 'Component',    'above': 'Section',       'col_name':'wCOMPinSECT'},
                4: {'below': 'Section',      'above': 'Technology',    'col_name':'wSECTinTECH'} }

        for r in range(len(df[lowest])):

            input[1] = df[lowest].loc[r,'Subcomponent']
            input[2] = df['Subcomponent'].query(f'Subcomponent == "{input[1]}"')['Component'].values[0]
            input[3] = df['Component'].query(f'Component == "{input[2]}"')['Section'].values[0]
            input[4] = df['Section'].query(f'Section == "{input[3]}"')['Technology'].values[0]

            for a in [2,3,4]:
                below = add_col[a]['below']
                above = add_col[a]['above']


                df[lowest].loc[r, above] = input[a]
                df[lowest].loc[r, add_col[a]['col_name']] = df[below].query(f"{above} == '{input[a]}' & {below} == '{input[a-1]}'")['Weight'].values[0]

        df[lowest]['Coeff'] = df[lowest]['Weight']* df[lowest]['wSUBCinCOMP'] * df[lowest]['wCOMPinSECT'] * df[lowest]['wSECTinTECH']
        
        if lowest == 'Factor of production':
            res[t][level] = df[level].loc[:,(lowest,'Coeff')]

        else:
            res[t][level] = df[level].loc[:,(lowest,'Region','Coeff')]

        res[t][level].set_index([x for x in res[t][level].columns if x != 'Coeff'], inplace=True)


    print(t, res[t]['Factor of production'].sum() + res[t]['Commodity'].sum())

# Check sum to 1
#%%

RES = {}

for e in ['Commodity', 'Factor of production']:
    if e != 'Factor of production':
        df1 = res['PV'][e].groupby([e,'Region']).sum()
        df2 = res['Offshore wind'][e].groupby([e,'Region']).sum()
        df3 = res['Onshore wind'][e].groupby([e,'Region']).sum()
        
        RES[e] = pd.concat([df1, df2, df3], axis=1).fillna(0)
        RES[e].columns = ['PV', 'Offshore wind', 'Onshore wind']

    else:
        df1 = res['PV'][e].groupby([e]).sum()
        df2 = res['Offshore wind'][e].groupby([e]).sum()
        df3 = res['Onshore wind'][e].groupby([e]).sum()

        RES[e] = pd.concat([df1, df2, df3], axis=1).fillna(0)
        RES[e].columns = ['PV', 'Offshore wind', 'Onshore wind']

# %% Now write on excel
# Commodity input
from openpyxl import load_workbook
path =r'C:\Users\nicog\Desktop\Nicolò\GitHub\GreenTechs'
sheet_names = ['input_from', 'Factor of production']
workbook = load_workbook(filename=path+'\Add Sectors/new_activities.xlsx')



commodity_region_grouped = res[t]['Commodity'].groupby(['Commodity', 'Region']).sum()
# Open the existing Excel file
# Select the worksheet "input_from"
for s in sheet_names:
    worksheet = workbook[s]

    if s != 'Factor of production':
        g = 'Commodity'
        # Write the contents of res['Commodity'] to the worksheet starting from row 4
        for index_c, row_c in RES[g].iterrows():
            print(index_c)
            pos_c = RES[g].index.get_loc(index_c)
            worksheet.cell(row=pos_c+4, column=1, value=index_c[1])
            worksheet.cell(row=pos_c+4, column=2, value='Commodity')
            worksheet.cell(row=pos_c+4, column=3, value=index_c[0])
            for t in tech_dict.keys():
                worksheet.cell(row=pos_c+4, column=tech_dict[t]['column_pos'], value=row_c.loc[t])
    else:
# Value added input
        fop_region_grouped = res[t]['Factor of production'].groupby(['Factor of production']).sum()
        worksheet = workbook['Factor of production']

        for index_f, row_f in RES[s].iterrows():
            print(index_f)
            pos_f = RES[s].index.get_loc(index_f)
            worksheet.cell(row=pos_f+4, column=1, value=index_f)
            for t in tech_dict.keys():
                worksheet.cell(row=pos_f+4, column=tech_dict[t]['column_pos']-2, value=row_f.loc[t])

# Save the changes to the file
workbook.save(filename=path+'\Add Sectors/new_activities.xlsx')
# %%
